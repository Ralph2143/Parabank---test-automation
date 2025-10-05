# parabank_bulk_payments_excel.py
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time, random, os
from openpyxl import Workbook, load_workbook
from login_helper import parabank_login  # ✅ Import reusable login

# ---- Config ----
CHROMEDRIVER_PATH = "chromedriver.exe"
BASE_URL = "https://parabank.parasoft.com/parabank/index.htm"
WAIT_SECONDS = 10
NUM_PAYMENTS = 100
EXCEL_FILE = "payments.xlsx"

first_names = ["John","Emma","Michael","Sophia","David","Olivia","James","Ava"]
last_names  = ["Smith","Johnson","Williams","Brown","Jones","Garcia","Miller","Davis"]
address     = ["Willowbrook","Stonehaven","Greenfield","Oakridge","Silvermere","Thornbury","Rivermist"]
city        = ["Eastbridge","Silverpoint","Grandchester","Ironvale","Brookhaven","Ashwick","Rivergate"]
state       = ["Silvermont","Westoria","Ashlandia","Crestwood"]
zip_code    = [3100,2300,3400,4555,6000,5560,3245]
phone_number= [33232145100,23546456400,34745745700,4557678678675,609679769600,5566796796790,3267967969645]
account_number = 2131231231
amounts     = [3100,2300,3400,4555,6000,5560,3245,8987,32,342,34234,34,3241,34161,6774,1212]

# ---- Excel Setup ----
if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Payment#", "Payee Name", "Amount", "Timestamp"])

def log_payment(i, name, amount):
    ws.append([i, name, amount, time.strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(EXCEL_FILE)

# ---- WebDriver ----
service = Service(executable_path=CHROMEDRIVER_PATH)
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)
wait = WebDriverWait(driver, WAIT_SECONDS)

try:
    print("[INFO] Opening Parabank...")
    parabank_login(driver, wait, BASE_URL)  # ✅ Login handled by helper

    for i in range(1, NUM_PAYMENTS + 1):
        print(f"[INFO] Starting payment #{i}...")
        bill_pay_link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Bill Pay")))
        bill_pay_link.click()
        wait.until(EC.presence_of_element_located((By.NAME, "payee.name")))

        random_name = f"{random.choice(first_names)} {random.choice(last_names)}"
        random_amount = random.choice(amounts)

        driver.find_element(By.NAME,"payee.name").send_keys(random_name)
        driver.find_element(By.NAME,"payee.address.street").send_keys(random.choice(address))
        driver.find_element(By.NAME,"payee.address.city").send_keys(random.choice(city))
        driver.find_element(By.NAME,"payee.address.state").send_keys(random.choice(state))
        driver.find_element(By.NAME,"payee.address.zipCode").send_keys(str(random.choice(zip_code)))
        driver.find_element(By.NAME,"payee.phoneNumber").send_keys(str(random.choice(phone_number)))
        driver.find_element(By.NAME,"payee.accountNumber").send_keys(str(account_number))
        driver.find_element(By.NAME,"verifyAccount").send_keys(str(account_number))
        driver.find_element(By.NAME,"amount").send_keys(str(random_amount))

        payment_btn = driver.find_element(By.CSS_SELECTOR,"input.button[value='Send Payment']")
        time.sleep(1)
        payment_btn.click()
        time.sleep(1)

        print(f"✅ Payment #{i} submitted.")
        log_payment(i, random_name, random_amount)

    print("🎉 All bulk payments completed and logged to Excel.")

except Exception as e:
    print("❌ Error occurred:", e)

finally:
    driver.quit()
    wb.save(EXCEL_FILE)
